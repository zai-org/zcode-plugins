#!/usr/bin/env python3
"""Recalculate an .xlsx workbook and report formula errors as JSON.

    python3 recalc.py <path_to_xlsx> [timeout_seconds] [--write-back]

openpyxl writes formula strings without evaluating them, so a freshly built
workbook has no cached values and no way to know whether its formulas compute
anything. This drives headless LibreOffice with recalc-on-load forced, then scans
for Excel error values.

**The caller's file is not modified by default.** The workbook is copied to a
temporary directory, recalculated there, and scanned there. An earlier version
copied LibreOffice's output back over the original with no backup, which is worse
than it sounds: LibreOffice's OOXML writer is not lossless, and the things it can
drop include cell comments — which is exactly where the model-building skills put
their source citations. Pass ``--write-back`` if you genuinely want the
recalculated file; a ``.bak`` is written first.

Exit codes (an earlier version returned 0 for every outcome, so ``&&`` chains
silently treated a workbook full of #REF! as a pass):

    0  success              every formula evaluated, no errors
    1  failed               usage error, missing file, timeout, LibreOffice error
    2  errors_found         formulas evaluated, at least one Excel error
    3  recalc_unavailable   LibreOffice missing; static lint only, NOT a pass

Output JSON (stdout):
  {
    "status": "success" | "errors_found" | "recalc_unavailable" | "failed",
    "total_errors": 0,
    "total_formulas": 42,
    "error_summary": {"#REF!": {"count": 2, "locations": ["DCF!B25", ...]}}
  }
"""
import json
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

EXCEL_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
MAX_LOCATIONS = 20  # cap per error type so output stays readable

EXIT_OK = 0
EXIT_FAILED = 1
EXIT_ERRORS = 2
EXIT_UNAVAILABLE = 3

# LibreOffice user-profile snippet: recalculate OOXML/ODF formulas on load
# without prompting (OOXMLRecalcMode / ODFRecalcMode 0 = always).
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry"
           xmlns:xs="http://www.w3.org/2001/XMLSchema">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
</oor:items>
"""


def emit(payload: dict, code: int) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    sys.exit(code)


def fail(msg: str) -> None:
    emit({"status": "failed", "error": msg}, EXIT_FAILED)


def find_soffice() -> str | None:
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def libreoffice_recalc(soffice: str, xlsx: Path, workdir: Path, timeout: float) -> Path:
    """Recalculate into `workdir` and return the recalculated file's path."""
    profile = workdir / "profile"
    (profile / "user").mkdir(parents=True)
    (profile / "user" / "registrymodifications.xcu").write_text(RECALC_XCU, encoding="utf-8")
    outdir = workdir / "out"
    outdir.mkdir()
    cmd = [
        soffice,
        f"-env:UserInstallation=file://{profile}",
        "--headless", "--norestore",
        "--convert-to", "xlsx",
        "--outdir", str(outdir),
        str(xlsx),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    converted = outdir / (xlsx.stem + ".xlsx")
    if proc.returncode != 0 or not converted.exists():
        raise RuntimeError(
            f"LibreOffice conversion failed (rc={proc.returncode}): "
            f"{(proc.stderr or proc.stdout).strip()[:500]}"
        )
    return converted


def scan_workbook(xlsx: Path) -> tuple[int, dict]:
    """Return (total_formulas, error_summary) from a recalculated workbook.

    Errors are read from cached values, so this is only meaningful on a file a
    spreadsheet engine has actually evaluated.
    """
    import openpyxl

    wb_f = openpyxl.load_workbook(xlsx, data_only=False)
    wb_v = openpyxl.load_workbook(xlsx, data_only=True)
    total_formulas = 0
    summary: dict[str, dict] = {}
    for ws_f, ws_v in zip(wb_f.worksheets, wb_v.worksheets):
        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            for cell_f, cell_v in zip(row_f, row_v):
                if cell_f.data_type == "f":
                    total_formulas += 1
                val = cell_v.value
                if isinstance(val, str) and val in EXCEL_ERRORS:
                    entry = summary.setdefault(val, {"count": 0, "locations": []})
                    entry["count"] += 1
                    if len(entry["locations"]) < MAX_LOCATIONS:
                        entry["locations"].append(f"{ws_f.title}!{cell_f.coordinate}")
    return total_formulas, summary


def static_lint(xlsx: Path) -> dict:
    """No-evaluation fallback: flag formula references to nonexistent sheets."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=False)
    sheets = set(wb.sheetnames)
    total_formulas = 0
    bad_refs: list[str] = []
    ref_pat = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_一-鿿]+))!")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                total_formulas += 1
                for m in ref_pat.finditer(cell.value):
                    name = m.group(1) or m.group(2)
                    if name not in sheets and len(bad_refs) < MAX_LOCATIONS:
                        bad_refs.append(f"{ws.title}!{cell.coordinate} -> '{name}'")
    result = {
        "status": "recalc_unavailable",
        "total_errors": len(bad_refs),
        "total_formulas": total_formulas,
        "note": ("LibreOffice (soffice) not found - formulas were NOT evaluated. "
                 "Static lint only, which is NOT a pass. Install LibreOffice for a "
                 "full check, or run the substitute checks in xlsx-author's SKILL.md."),
    }
    if bad_refs:
        result["error_summary"] = {
            "BROKEN_SHEET_REF": {"count": len(bad_refs), "locations": bad_refs}
        }
    return result


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    unknown = flags - {"--write-back"}
    if unknown:
        fail(f"unknown option(s): {', '.join(sorted(unknown))}")
    if not args:
        fail("usage: recalc.py <path_to_xlsx> [timeout_seconds] [--write-back]")

    xlsx = Path(args[0])
    try:
        timeout = float(args[1]) if len(args) > 1 else 30.0
    except ValueError:
        fail(f"timeout must be a number, got {args[1]!r}")
    if not xlsx.is_file():
        fail(f"file not found: {xlsx}")

    soffice = find_soffice()
    if soffice is None:
        emit(static_lint(xlsx), EXIT_UNAVAILABLE)

    with tempfile.TemporaryDirectory(prefix="recalc-") as tmp:
        workdir = Path(tmp)
        # Recalculate a copy so the caller's file — and its cell comments, which
        # carry source citations — survive LibreOffice's lossy round-trip.
        staged = workdir / xlsx.name
        shutil.copyfile(xlsx, staged)
        try:
            recalculated = libreoffice_recalc(soffice, staged, workdir, timeout)
        except subprocess.TimeoutExpired:
            fail(f"LibreOffice recalculation timed out after {timeout}s")
        except Exception as exc:  # noqa: BLE001
            fail(str(exc))

        total_formulas, summary = scan_workbook(recalculated)
        backup = None
        if "--write-back" in flags:
            backup = xlsx.with_suffix(xlsx.suffix + ".bak")
            shutil.copyfile(xlsx, backup)
            shutil.copyfile(recalculated, xlsx)

    total_errors = sum(e["count"] for e in summary.values())
    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": total_formulas,
    }
    if summary:
        result["error_summary"] = summary
    if backup:
        result["wrote_back"] = True
        result["backup"] = str(backup)
    emit(result, EXIT_OK if total_errors == 0 else EXIT_ERRORS)


if __name__ == "__main__":
    main()
